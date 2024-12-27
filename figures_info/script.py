# !/usr/bin/env python
# -*- coding: utf-8 -*-
# @Time    : 2024/11/7 14:52
# @Author  : incpink Liu
# @File    : script.py
import math
import csv
import os

import numpy as np
import pandas as pd
import seaborn as sns
import openpyxl
import matplotlib
import matplotlib.pyplot as plt
import matplotlib.gridspec as gridspec
from matplotlib.pyplot import tick_params
from mpl_toolkits.axes_grid1.inset_locator import mark_inset
from mpl_toolkits.axes_grid1.inset_locator import inset_axes


def figure_2():
    # Overall config.
    figure, ((ax_1, ax_2), (ax_3, ax_4)) = plt.subplots(2, 2, figsize=(12, 12))
    ax_3_ins = inset_axes(
        ax_3,
        width="40%",
        height="40%",
        loc="lower left",
        bbox_to_anchor=(0.13, 0.07, 0.7, 1.3),
        bbox_transform=ax_3.transAxes
    )
    mark_inset(ax_3, ax_3_ins, loc1=2, loc2=1, fc="none", ec="k", lw=1, ls="--")
    sns.set_theme(style="white", font="DejaVu Sans")
    sns.set_context("paper", font_scale=1.2)
    palette = "Set2"

    # Figure_2-A data.
    data_A = pd.DataFrame()
    for case in os.scandir("./input_data/Figure2A_GC_content"):
        csv_data = pd.read_csv(case.path)
        data_A = pd.concat([data_A, csv_data], axis=0)
    data_A["GC content"] = data_A["GC content"].multiply(100)

    # Figure_2-B data.
    data_B_path = "./input_data/Figure2B_Homopolymers/homopolymer_length.xlsx"
    data_B = pd.read_excel(data_B_path)
    file_names = [file_name for file_name in data_B.columns[1:].tolist()]

    # Figure_2-C and figure_2-D data.
    data_C_path = "./input_data/Figure2C_Average_recovery_rate/average_recovery_rate.xlsx"
    data_D_path = "./input_data/Figure2D_Coefficient_variation/coefficient_variation.xlsx"
    data_C, data_D = pd.read_excel(data_C_path), pd.read_excel(data_D_path)
    avg_values = data_C.values[:, 1:].astype(np.float32)
    cov_values = data_D.values[:, 1:].astype(np.float32)
    error_rates = [float(error_rate[: -1]) for error_rate in data_D.columns[1:].tolist()]

    # -------------------------------Figure_2-A-------------------------------
    sns.violinplot(
        x="file name",
        y="GC content",
        data=data_A,
        hue="file name",
        bw_method=0.15,
        density_norm="width",
        palette=palette,
        legend=False,
        ax=ax_1,
        linewidth=1.5,
    )
    ax_1.set_ylabel("% GC content")
    ax_1.set_xticks(ticks=np.arange(6), labels=file_names, rotation=15)
    ax_1.set_xlabel("")
    ax_1.text(-0.13, 1.05, "(A)", ha="left", va="top", transform=ax_1.transAxes)

    ax_1.spines["top"].set_linewidth(1.5)
    ax_1.spines["right"].set_linewidth(1.5)
    ax_1.spines["bottom"].set_linewidth(1.5)
    ax_1.spines["left"].set_linewidth(1.5)

    # -------------------------------Figure_2-B---------------------------------
    sns.boxplot(
        data=data_B.values,
        palette=palette,
        ax=ax_2,
        linewidth=1.5,
    )
    ax_2.set_ylabel("Homopolymers")
    ax_2.set_xticks(ticks=np.arange(6), labels=file_names, rotation=15)
    ax_2.text(-0.13, 1.05, "(B)", ha="left", va="top", transform=ax_2.transAxes)

    ax_2.spines["top"].set_linewidth(1.5)
    ax_2.spines["right"].set_linewidth(1.5)
    ax_2.spines["bottom"].set_linewidth(1.5)
    ax_2.spines["left"].set_linewidth(1.5)

    # ----------------------------Figure_2-C & D-------------------------------
    colors = sns.color_palette(palette)
    for i in range(len(avg_values)):
        ax_3.plot(error_rates, avg_values[i], marker="o", markersize=12, color=colors[i], markeredgecolor="white")
        ax_3_ins.plot(error_rates[: 3], avg_values[i][: 3], marker="o", markersize=10, color=colors[i], markeredgecolor="white")
        ax_4.plot(error_rates, cov_values[i], marker="o", markersize=12, color=colors[i], markeredgecolor="white")

    ax_3.set_xlabel("% Error rate")
    ax_3.set_ylabel("% Avg of recovery rate")
    ax_3.set_yticks(np.arange(78, 102, 2))
    ax_3.set_ylim(93, 101)
    ax_3.text(-0.13, 1.05, "(C)", ha="left", va="top", transform=ax_3.transAxes)

    ax_3_ins.set_xlim(0, 0.25)
    ax_3_ins.set_ylim(99.4, 100.05)
    ax_3_ins.set_xticks(np.arange(0, 0.25, 0.1))
    ax_3_ins.set_yticks(np.arange(99.4, 100.05, 0.05))

    ax_3_ins.spines["top"].set_linewidth(1.5)
    ax_3_ins.spines["right"].set_linewidth(1.5)
    ax_3_ins.spines["bottom"].set_linewidth(1.5)
    ax_3_ins.spines["left"].set_linewidth(1.5)

    ax_3.spines["top"].set_linewidth(1.5)
    ax_3.spines["right"].set_linewidth(1.5)
    ax_3.spines["bottom"].set_linewidth(1.5)
    ax_3.spines["left"].set_linewidth(1.5)

    ax_4.set_xlabel("% Error rate")
    ax_4.set_ylabel("% Cov of recovery rate")
    ax_4.text(-0.13, 1.05, "(D)", ha="left", va="top", transform=ax_4.transAxes)

    ax_4.spines["top"].set_linewidth(1.5)
    ax_4.spines["right"].set_linewidth(1.5)
    ax_4.spines["bottom"].set_linewidth(1.5)
    ax_4.spines["left"].set_linewidth(1.5)

    fig2_path = "./generated_figure/Figure2.svg"
    plt.savefig(fig2_path)
    plt.show()


def figure_3_without_subplot_A():
    # Overall config.
    fig = plt.figure(figsize=(16, 18))
    gs = gridspec.GridSpec(40, 20)
    sns.set_theme(style="white", font="DejaVu Sans")
    sns.set_context("paper", font_scale=1.2)

    # Figure_3-B & C data.
    sequence_groups = ["A", "B", "C", "D", "E", "F", "G", "H", "I", "J", "K", "L", "M", "N", "O"]
    data_B_C_path = "./input_data/Figure3B&C_Error_count_record/errors_count_record.xlsx"
    xlsx = openpyxl.load_workbook(data_B_C_path)
    sheet = xlsx["Sheet1"]

    insertions_ratio, deletions_ratio, substitutions_ratio, CM_sub_ratio = [], [], [], []
    insertions_cnt, deletions_cnt, substitutions_cnt, CM_sub_cnt = [], [], [], []

    for i in range(1, 31, 2):
        ins_cnt = int(sheet[f"R{i * 3}"].value) + int(sheet[f"R{(i + 1) * 3 }"].value)
        del_cnt = int(sheet[f"R{i * 3 + 1}"].value) + int(sheet[f"R{(i + 1) * 3 + 1}"].value)
        sub_cnt = int(sheet[f"R{i * 3 + 2}"].value.split("/")[-1]) + int(sheet[f"R{(i + 1) * 3 + 2}"].value.split("/")[-1])
        CM_cnt = int(sheet[f"R{i * 3 + 2}"].value.split("/")[0]) + int(sheet[f"R{(i + 1) * 3 + 2}"].value.split("/")[0])
        summary = ins_cnt + del_cnt + sub_cnt

        sub_ratio = (sub_cnt/ summary) * 100
        ins_ratio = (ins_cnt / summary) * 100
        del_ratio = 100 - sub_ratio - ins_ratio
        CM_ratio = (CM_cnt / sub_cnt) * 100

        insertions_ratio.append(ins_ratio)
        deletions_ratio.append(del_ratio)
        substitutions_ratio.append(sub_ratio)
        CM_sub_ratio.append(CM_ratio)

        insertions_cnt.append(ins_cnt)
        deletions_cnt.append(del_cnt)
        substitutions_cnt.append(sub_cnt)
        CM_sub_cnt.append(CM_cnt)

    single_CM_ratio = []
    for m in range(1, 31):
        for n in [chr(x) for x in range(67, 82)]:
            try:
                single_CM_cnt, single_sub_cnt = [int(y) for y in sheet[f"{n}{m * 3 + 2}"].value.split("/")]
            except AttributeError:
                break
            else:
                single_CM_ratio.append((single_CM_cnt / single_sub_cnt) * 100) if single_sub_cnt != 0 else single_CM_ratio.append(0)

    data_B = [substitutions_ratio[:: -1], deletions_ratio[:: -1], insertions_ratio[:: -1]]
    data_C = np.asarray(single_CM_ratio).reshape(27, 15).T

    # Figure_3-D data.
    location_error_without, location_error_with = [], []
    data_D_path = "./input_data/Figure3D_Error_location_record/errors_location_record.xlsx"
    wb = openpyxl.load_workbook(data_D_path)
    ws_without = wb["without_ref"]
    ws_with = wb["with_ref"]

    for col in range(2, 102):
        summation_without, summation_with = 0, 0
        for row in range(2, 407):
            summation_without += ws_without.cell(row, col).value
            summation_with += ws_with.cell(row, col).value
        location_error_without.append((summation_without / 405) * 100)
        location_error_with.append((summation_with / 405) * 100)

    data_D = [list(range(1, 101)), location_error_without, location_error_with]

    # -------------------------------Figure_3-A-------------------------------
    # Blank subplot.
    ax_0 = fig.add_subplot(gs[0: 16, :])
    ax_0.text(-0.057, 1.05, "(A)", ha="left", va="top", transform=ax_0.transAxes)
    ax_0.axis("off")

    # -------------------------------Figure_3-B-------------------------------
    ax_1 = fig.add_subplot(gs[16: 27, 0: 7])
    ax_1.tick_params(direction="out", bottom=True, left=True)

    color = [sns.color_palette()[-1], sns.color_palette()[-2], sns.color_palette()[-3]]

    for d in range(len(data_B)):
        ax_1.barh(sequence_groups[:: -1], data_B[d], left=np.sum(data_B[:d], axis=0), color=color[d], edgecolor="black", height=0.75)

    ax_1.legend(ncols=3, loc="upper center", labels=["sub.", "del.", "ins."], bbox_to_anchor=(0.5, 1.12), prop={"size": 12})

    ax_1.set_yticks(ticks=np.arange(15), labels=sequence_groups[:: -1])
    ax_1.set_ylabel("Group", size=12)
    bottom, top = ax_1.get_ylim()
    ax_1.set_ylim(bottom + 0.5, top - 0.6)

    ax_1.set_xlabel("% Ratio", size=12)
    ax_1.set_xlim(0, 100)

    ax_1.text(-0.16, 1.1, "(B)", ha="left", va="top", transform=ax_1.transAxes)

    ax_1.spines["top"].set_visible(False)
    ax_1.spines["right"].set_visible(False)
    ax_1.spines["left"].set_visible(False)
    ax_1.spines["bottom"].set_linewidth(2)

    # -------------------------------Figure_3-C-------------------------------
    ax_2 = fig.add_subplot(gs[16: 27, 8: ])
    ax_2.tick_params(direction="out", bottom=True, left=True)

    sns.heatmap(
        data_C,
        cmap="GnBu",
        ax=ax_2,
        cbar=True,
        cbar_kws={"label": "% Ratio of sub. between C and M", "pad": 0.02},
        yticklabels=sequence_groups,
        linewidths=0.75,
        xticklabels=4,
    )

    ax_2.set_xticklabels(labels=list(range(0, 27, 4)), rotation=0)
    ax_2.set_xlabel("Oligo in group", size=12)

    ax_2.set_yticklabels(labels=sequence_groups, rotation=0)
    ax_2.text(-0.07, 1.1, "(C)", ha="left", va="top", transform=ax_2.transAxes)

    ax_2.spines["bottom"].set_linewidth(2)

    # -------------------------------Figure_3-D-------------------------------
    ax_3 = fig.add_subplot(gs[29:, :])
    ax_3.tick_params(direction="out", bottom=True, left=True)

    ax_3.plot(data_D[0], data_D[1], color=sns.color_palette()[3], marker="o", markersize=5)
    ax_3.plot(data_D[0], data_D[2], color=sns.color_palette()[4], marker="o", markersize=5)

    ax_3.set_xlabel("Base position in oligo", size=12)
    ax_3.set_xticks(ticks=np.arange(0, 101, 10))

    ax_3.set_ylabel("% Error rate", size=12)
    ax_3.set_ylim(0, 16)

    ax_3.legend(["without reference", "with reference"], loc="upper right", prop={"size": 12}, bbox_to_anchor=(0.9, 0.95))

    ax_3.text(-0.06, 1.05, "(D)", ha="left", va="top", transform=ax_3.transAxes)

    ax_3.spines["top"].set_visible(False)
    ax_3.spines["right"].set_visible(False)
    ax_3.spines["bottom"].set_linewidth(2)
    ax_3.spines["left"].set_linewidth(2)

    fig3_path = "./generated_figure/Figure3_without_subplot_A.svg"
    plt.savefig(fig3_path)
    plt.show()


def supplementary_figure_1():
    x = np.linspace(0, 10, 100)
    y = np.log2(x + 1)

    fig, ax = plt.subplots()

    ax.plot(x[: 30], y[: 30], linestyle="--", color="royalblue")
    ax.plot(x[30:], y[30:], linestyle="-", label="H = log$_2$(N + 1)", color="royalblue")

    ax.set_xlabel("value of N", fontsize=10)
    ax.set_ylabel("limit of information density / bits per base", fontsize=10)
    ax.legend(loc="upper left")

    x_ = [3, 7]
    y_ = [2.0, 3.0]
    for i in range(len(x_)):
        ax.scatter(x_[i], y_[i], color="black", s=50)
        ax.text(x_[i] + 0.3, y_[i] - 0.1, f"({x_[i]}, {y_[i]})", fontsize=11, color="black")

    plt.axhline(y=2, color="black", linestyle="-.", xmin=0, xmax=0.32)
    plt.axhline(y=3, color="black", linestyle="-.", xmin=0, xmax=0.68)
    plt.axvline(x=3, color="black", linestyle="-.", ymin=0, ymax=0.57)
    plt.axvline(x=7, color="black", linestyle="-.", ymin=0, ymax=0.83)

    figs1_path = "./generated_figure/FigureS1.svg"
    plt.savefig(figs1_path)
    plt.show()


def supplementary_figure_2():
    cases = ["Goldman.pdf", "poem.txt", "program.py", "rain.wav", "speak.mp4", "sunflower.bmp"]
    fig, ((ax_1, ax_2), (ax_3, ax_4), (ax_5, ax_6)) = plt.subplots(3, 2, figsize=(12, 22), gridspec_kw={"left": 0.055, "right": 0.89, "wspace": 0.15, "hspace": 0.01})
    cmap = "GnBu_r"

    data_set = []
    for case in cases:
        case_path = f"./input_data/FigureS2/{case}-recovery_rates.xlsx"
        data = pd.read_excel(case_path)
        recovery_rates = data.values[:, 1:].astype(np.float32).T
        error_rates = [error_rate[: -1] for error_rate in data.columns[1:].tolist()]
        test_times = data.values[:, 0].tolist()
        data_set.append((recovery_rates, error_rates, test_times))

    image_1 = heatmap(
        data=data_set[0][0],
        row_labels=data_set[0][1],
        column_labels=data_set[0][2],
        x_label=cases[0],
        y_label="% Error rate",
        ax=ax_1,
        need_xticks=True,
        need_yticks=True,
        need_cbar=False,
        cmap=cmap,
    )
    annotate_heatmap(image_1, valfmt="{x:.2f}", size=9)
    ax_1.xaxis.tick_top()
    ax_1.set_xticklabels(ax_1.get_xticklabels(), rotation=15)

    image_2 = heatmap(
        data=data_set[1][0],
        row_labels=data_set[1][1],
        column_labels=data_set[1][2],
        x_label=cases[1],
        ax=ax_2,
        need_xticks=True,
        need_yticks=False,
        need_cbar=False,
        cmap=cmap,
    )
    annotate_heatmap(image_2, valfmt="{x:.2f}", size=9)
    ax_2.xaxis.tick_top()
    ax_2.set_xticklabels(ax_2.get_xticklabels(), rotation=15)

    image_3 = heatmap(
        data=data_set[2][0],
        row_labels=data_set[2][1],
        column_labels=data_set[2][2],
        x_label=cases[2],
        y_label="% Error rate",
        ax=ax_3,
        need_xticks=False,
        need_yticks=True,
        need_cbar=False,
        cmap=cmap,
    )
    annotate_heatmap(image_3, valfmt="{x:.2f}", size=9)

    image_4 = heatmap(
        data=data_set[3][0],
        row_labels=data_set[3][1],
        column_labels=data_set[3][2],
        x_label=cases[3],
        ax=ax_4,
        need_xticks=False,
        need_yticks=False,
        need_cbar=False,
        cmap=cmap,
    )
    annotate_heatmap(image_4, valfmt="{x:.2f}", size=9)

    image_5 = heatmap(
        data=data_set[4][0],
        row_labels=data_set[4][1],
        column_labels=data_set[4][2],
        x_label=cases[4],
        y_label="% Error rate",
        ax=ax_5,
        need_xticks=False,
        need_yticks=True,
        need_cbar=False,
        cmap=cmap,
    )
    annotate_heatmap(image_5, valfmt="{x:.2f}", size=9)

    image_6 = heatmap(
        data=data_set[5][0],
        row_labels=data_set[5][1],
        column_labels=data_set[5][2],
        x_label=cases[5],
        ax=ax_6,
        need_xticks=False,
        need_yticks=False,
        need_cbar=False,
        cmap=cmap,
    )
    annotate_heatmap(image_6, valfmt="{x:.2f}", size=9)

    l = 0.915
    b = 0.122
    w = 0.03
    h = 0.746
    rect = [l, b, w, h]
    cbar_ax = fig.add_axes(rect)
    cbar = fig.colorbar(image_6, cax=cbar_ax)

    cbar.set_label("% Data recovery rate", rotation=90, fontsize=12)

    figs2_path = "./generated_figure/FigureS2.svg"
    plt.savefig(figs2_path)
    plt.show()


def heatmap(
        data: np.ndarray,
        row_labels: list = None,
        column_labels: list = None,
        x_label: str = None,
        y_label: str = None,
        ax=None,
        cbar_kw: dict = None,
        need_xticks: bool = True,
        need_yticks: bool = True,
        need_cbar: bool = True,
        cbarlabel: str = "",
        **kwargs
):
    if ax is None:
        ax = plt.gca()

    if cbar_kw is None:
        cbar_kw = {}

    image = ax.imshow(data, **kwargs)

    if need_xticks:
        ax.set_xticks(np.arange(data.shape[1]), labels=column_labels, fontsize=10)
    else:
        ax.set_xticks([])

    if need_yticks:
        ax.set_yticks(np.arange(data.shape[0]), labels=row_labels, fontsize=11)
    else:
        ax.set_yticks([])
    ax.set_xlabel(xlabel=x_label, fontsize=12) if x_label is not None else None
    ax.set_ylabel(ylabel=y_label, fontsize=12) if y_label is not None else None

    ax.tick_params(top=False, bottom=True, labeltop=False, labelbottom=True)

    plt.setp(ax.get_xticklabels(), rotation=15)

    ax.spines[:].set_visible(False)

    ax.set_xticks(np.arange(data.shape[1] + 1) - 0.5, minor=True)
    ax.set_yticks(np.arange(data.shape[0] + 1) - 0.5, minor=True)
    ax.grid(which="minor", color="white", linestyle="-", linewidth=3)
    ax.tick_params(which="minor", bottom=False, left=False)

    if need_cbar:
        cbar = ax.figure.colorbar(image, ax=ax, **cbar_kw)
        cbar.ax.set_ylabel(cbarlabel, rotation=-90, va="bottom")

        return image, cbar

    return image


def annotate_heatmap(
        image,
        data=None,
        valfmt="{x:.2f}",
        textcolors=("black", "white"),
        threshold=None,
        **textkw
) -> list:
    if not isinstance(data, (list, np.ndarray)):
        data = image.get_array()

    if threshold is not None:
        threshold = image.norm(threshold)
    else:
        threshold = image.norm(data.max()) / 2

    kw = dict(horizontalalignment="center", verticalalignment="center")
    kw.update(textkw)

    if isinstance(valfmt, str):
        valfmt = matplotlib.ticker.StrMethodFormatter(valfmt)

    texts = []
    for i in range(data.shape[0]):
        for j in range(data.shape[1]):
            kw.update(color=textcolors[int(image.norm(data[i, j]) < threshold)])
            text = image.axes.text(j, i, valfmt(data[i, j], None), **kw)
            texts.append(text)

    return texts


def supplementary_figure_3():
    file_names = ["Goldman.pdf", "poem.txt", "program.py", "rain.wav", "speak.mp4", "sunflower.bmp"]
    file_sizes = [os.path.getsize(f"../test_files/for_simulation_evaluation/{file_name}") for file_name in file_names]

    GC_xmax_xmin = []
    for case in os.scandir(r"./input_data/FigureS3/GC_content"):
        csv_data = pd.read_csv(case.path)
        GC_xmax_xmin.append(csv_data["GC content"].multiply(100).max() - csv_data["GC content"].multiply(100).min())

    homopolymer_max_len = pd.read_excel(r"./input_data/FigureS3/Homopolymer/homopolymer_length.xlsx").values[1][1:]

    all_data = sorted(list(zip(file_names, file_sizes, GC_xmax_xmin, homopolymer_max_len)), key=lambda x: x[1])
    sorted_file_names = [data[0] for data in all_data]
    sorted_file_sizes = [data[1] for data in all_data]
    sorted_GC_xmax_xmin = [data[2] for data in all_data]
    sorted_homopolymer_max_len = [data[3] for data in all_data]

    palette = sns.color_palette("Set2")
    c = [palette[1], palette[2], palette[0], palette[4], palette[3], palette[5]]

    figure, (ax_1, ax_2) = plt.subplots(2, 1, figsize=(6, 12))

    m_GC, b_GC = np.polyfit(np.log10(sorted_file_sizes), sorted_GC_xmax_xmin, 1)
    m_homopolymer, b_homopolymer = np.polyfit(np.log10(sorted_file_sizes), sorted_homopolymer_max_len, 1)
    file_size_fit = np.linspace(min(sorted_file_sizes), max(sorted_file_sizes), 100)
    GC_fit = m_GC * np.log10(file_size_fit) + b_GC
    homopolymer_fit = m_homopolymer * np.log10(file_size_fit) + b_homopolymer

    ax_1.set_xscale("log")
    for i in range(len(sorted_file_names)):
        ax_1.scatter(sorted_file_sizes[i], sorted_GC_xmax_xmin[i], marker="o", color=c[i], s=150)
    ax_1.set_xticks([1000, 10000, 100000, 1000000], labels=["1 KB", "10 KB", "100 KB", "1 MB"], rotation=15)
    ax_1.plot(file_size_fit, GC_fit, color="black", linestyle="--")
    ax_1.set_xlabel("File size")
    ax_1.set_ylabel("% Range of GC content")
    ax_1.legend(loc="upper left", labels=sorted_file_names)
    ax_1.text(-0.15, 1.05, "(A)", ha="left", va="top", transform=ax_1.transAxes)

    ax_2.set_xscale("log")
    for j in range(len(sorted_file_names)):
        ax_2.scatter(sorted_file_sizes[j], sorted_homopolymer_max_len[j], marker="^", color=c[j], s=150)
    ax_2.set_xticks([1000, 10000, 100000, 1000000], labels=["1 KB", "10 KB", "100 KB", "1 MB"], rotation=15)
    ax_2.plot(file_size_fit, homopolymer_fit, color="black", linestyle="-.")
    ax_2.set_xlabel("File size")
    ax_2.set_ylabel("Max length of homopolymer / nt")
    ax_2.legend(loc="upper left", labels=sorted_file_names)
    ax_1.text(-0.15, 1.05, "(B)", ha="left", va="top", transform=ax_2.transAxes)

    figs3_path = "./generated_figure/FigureS3.svg"
    plt.savefig(figs3_path)
    plt.show()


def supplementary_figure_6():
    loc_error_without_sub, loc_error_without_ins, loc_error_without_del, without_errors = [], [], [], []
    data_path = r"input_data/FigureS6/errors_location_record.xlsx"
    wb = openpyxl.load_workbook(data_path)
    ws_without_sub = wb["without_ref_sub"]
    ws_without_ins = wb["without_ref_ins"]
    ws_without_del = wb["without_ref_del"]

    for col in range(2, 102):
        summation_without_sub, summation_without_ins, summation_without_del, no_errors = 0, 0, 0, 0
        for row in range(2, 407):
            summation_without_sub += ws_without_sub.cell(row, col).value
            summation_without_ins += ws_without_ins.cell(row, col).value
            summation_without_del += ws_without_del.cell(row, col).value
        total = summation_without_sub + summation_without_ins + summation_without_del
        if total != 0:
            sub_ratio = (summation_without_sub / total) * 100
            ins_ratio = (summation_without_ins / total) * 100
            del_ratio = 100 - sub_ratio - ins_ratio
        else:
            sub_ratio, ins_ratio, del_ratio = 0, 0, 0
            no_errors = 100

        loc_error_without_sub.append(sub_ratio)
        loc_error_without_ins.append(ins_ratio)
        loc_error_without_del.append(del_ratio)
        without_errors.append(no_errors)

    data_A = [loc_error_without_sub[: 50], loc_error_without_del[: 50], loc_error_without_ins[: 50], without_errors[: 50]]
    data_B = [loc_error_without_sub[50:], loc_error_without_del[50:], loc_error_without_ins[50:], without_errors[50:]]

    gs = gridspec.GridSpec(40, 20)
    ax_1 = plt.subplot(gs[: 18, : 20])
    ax_2 = plt.subplot(gs[22: 40, : 20])

    ax_1.tick_params(direction="out", bottom=True, left=True)

    color = [sns.color_palette()[-1], sns.color_palette()[-2], sns.color_palette()[-3], sns.color_palette()[0]]

    for d in range(len(data_A)):
        ax_1.bar(list(range(1, 51)), data_A[d], bottom=np.sum(data_A[:d], axis=0), color=color[d], width=0.5, edgecolor="black")
        ax_2.bar(list(range(51, 101)), data_B[d], bottom=np.sum(data_B[:d], axis=0), color=color[d], width=0.5, edgecolor="black")

    ax_1.legend(ncols=4, loc="upper center", labels=["sub.", "del.", "ins.", "no_errors"], bbox_to_anchor=(0.5, 1.3), prop={"size": 10})
    ax_1.set_ylabel("% Ratio", size=10)
    ax_2.set_ylabel("% Ratio", size=10)
    ax_2.set_xlabel("Base position in oligo", size=10)

    figs5_path = "./generated_figure/FigureS6.svg"
    plt.savefig(figs5_path)
    plt.show()


if __name__ == "__main__":
    pass
